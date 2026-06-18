"""
回归测试：覆盖现有 test_*.py 漏掉的接口与边界场景

策略：每个测试都建独立测试用户，跑完自动清理。
"""
import io
import sys
import time
import pytest

sys.path.insert(0, '.')

from gerenjizhang.app import app
from gerenjizhang.db import Database, add_record, get_records
from gerenjizhang.utils.classifier import (
    clear_user_learned, invalidate_cache, classify as smart_classify
)


# ==================== Fixtures ====================
@pytest.fixture
def client():
    app.config['TESTING'] = True
    with app.test_client() as c:
        yield c


def _make_user():
    """注册新用户 + 拿 token + 返回 (client-headers, user_id)"""
    suffix = str(int(time.time() * 1000))[-8:]
    username = f'reg_{suffix}'

    with Database() as db:
        db.c.execute(
            'INSERT INTO users (username, password) VALUES (%s, %s)',
            (username, 'pbkdf2_sha256$dummy')
        )
        user_id = db.c.lastrowid
        db.conn.commit()

    from gerenjizhang.utils.decorators import set_token, remove_token
    token = f'reg_token_{suffix}'
    set_token(token, user_id)
    return username, token, user_id


@pytest.fixture
def reg_user():
    """注册一个测试用户，yield (auth_client, user_id)"""
    username, token, user_id = _make_user()

    class AuthClient:
        def __init__(self, c, token, user_id):
            self.c = c
            self.token = token
            self.user_id = user_id
            self.headers = {'Authorization': f'Bearer {token}'}

        def get(self, *a, **kw):
            kw.setdefault('headers', {}).update(self.headers)
            return self.c.get(*a, **kw)

        def post(self, *a, **kw):
            kw.setdefault('headers', {}).update(self.headers)
            return self.c.post(*a, **kw)

    client = app.test_client()
    ac = AuthClient(client, token, user_id)
    yield ac, user_id

    # 清理
    from gerenjizhang.utils.decorators import remove_token
    remove_token(token)
    clear_user_learned(user_id)
    with Database() as db:
        db.c.execute('DELETE FROM bill WHERE user_id=%s', (user_id,))
        db.c.execute('DELETE FROM budgets WHERE user_id=%s', (user_id,))
        db.c.execute('DELETE FROM user_classify_memory WHERE user_id=%s', (user_id,))
        db.c.execute('DELETE FROM import_history WHERE user_id=%s', (user_id,))
        db.c.execute('DELETE FROM tokens WHERE user_id=%s', (user_id,))
        db.c.execute('DELETE FROM users WHERE id=%s', (user_id,))
        db.conn.commit()


# ==================== /api/bill/<rid> GET 详情 ====================
class TestGetBillDetail:
    def test_get_own_record(self, reg_user):
        ac, uid = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮', 'note': '午饭'
        })
        rid = res.get_json()['data']['id']

        res = ac.get(f'/api/bill/{rid}')
        assert res.get_json()['code'] == 0
        d = res.get_json()['data']
        assert d['id'] == rid
        assert d['amount'] == 100
        assert d['category'] == '餐饮'
        assert d['note'] == '午饭'

    def test_get_other_user_record_404(self, client):
        """别人的记录：直接查 API 应该返回「记录不存在」而不是泄漏数据"""
        # 创建用户 A 并拿他的 id
        suffix = str(int(time.time() * 1000))[-8:]
        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'detail_a_{suffix}', 'x'))
            uid_a = db.c.lastrowid
            db.conn.commit()
        add_record('2026-06-01', 100, 'expense', '餐饮', '午饭', uid_a)
        recs = get_records(uid_a)
        rid = recs[0][0]

        # 创建用户 B
        _, token, uid_b = _make_user()
        from gerenjizhang.utils.decorators import set_token, remove_token
        # 用户 B 试图看 A 的记录
        res = client.get(f'/api/bill/{rid}', headers={'Authorization': f'Bearer {token}'})
        assert res.get_json()['code'] == 2001  # 记录不存在

        # 清理
        remove_token(token)
        with Database() as db:
            db.c.execute('DELETE FROM bill WHERE user_id IN (%s, %s)', (uid_a, uid_b))
            db.c.execute('DELETE FROM users WHERE id IN (%s, %s)', (uid_a, uid_b))
            db.conn.commit()

    def test_get_nonexistent_record(self, reg_user):
        ac, _ = reg_user
        res = ac.get('/api/bill/99999')
        assert res.get_json()['code'] == 2001


# ==================== 智能分类 API ====================
class TestClassifyAPI:
    def test_classify_known_brand(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/classify', json={'text': '美团外卖 28.5'})
        assert res.get_json()['code'] == 0
        d = res.get_json()['data']
        assert d['category'] == '餐饮'
        assert d['type'] == 'expense'

    def test_classify_empty_text_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/classify', json={'text': ''})
        assert res.get_json()['code'] == 6001

    def test_classify_whitespace_text_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/classify', json={'text': '   '})
        assert res.get_json()['code'] == 6001

    def test_classify_unknown_returns_null(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/classify', json={'text': '哈哈呵呵'})
        # 智能分类未识别 → code=0 但 data.category=None
        assert res.get_json()['code'] == 0
        assert res.get_json()['data']['category'] is None


# ==================== 分类记忆 API ====================
class TestClassifyMemoryAPI:
    def test_list_memory_empty(self, reg_user):
        ac, uid = reg_user
        res = ac.get('/api/bill/classify/memory')
        assert res.get_json()['code'] == 0
        assert res.get_json()['data'] == []

    def test_list_memory_after_edit_learns(self, reg_user):
        """回归：用户改分类 → 应该写入记忆 → 列表能看到"""
        ac, uid = reg_user
        # 添加一条
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '其他', 'note': '某某小店'
        })
        rid = res.get_json()['data']['id']
        # 编辑改成餐饮
        ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '餐饮', 'note': '某某小店'
        })

        res = ac.get('/api/bill/classify/memory')
        memories = res.get_json()['data']
        assert any(m['keyword'] == '某某小店' and m['category'] == '餐饮' for m in memories), \
            f"编辑后没学习到记忆: {memories}"

    def test_delete_memory(self, reg_user):
        ac, uid = reg_user
        # 加一条记忆
        from gerenjizhang.utils.classifier import remember_correction
        remember_correction(uid, '测试关键词', '餐饮', 'expense')

        res = ac.get('/api/bill/classify/memory')
        memories = res.get_json()['data']
        mid = next(m['id'] for m in memories if m['keyword'] == '测试关键词')

        res = ac.post(f'/api/bill/classify/memory/{mid}/delete') \
            if False else ac.c.delete(f'/api/bill/classify/memory/{mid}',
                                       headers=ac.headers)
        assert res.get_json()['code'] == 0

        # 再查应该没了
        res = ac.get('/api/bill/classify/memory')
        assert not any(m['id'] == mid for m in res.get_json()['data'])

    def test_delete_nonexistent_memory(self, reg_user):
        ac, _ = reg_user
        res = ac.c.delete('/api/bill/classify/memory/99999', headers=ac.headers)
        assert res.get_json()['code'] == 6002

    def test_clear_memory(self, reg_user):
        ac, uid = reg_user
        from gerenjizhang.utils.classifier import remember_correction
        remember_correction(uid, '记忆1', '餐饮', 'expense')
        remember_correction(uid, '记忆2', '购物', 'expense')

        res = ac.post('/api/bill/classify/memory/clear')
        assert res.get_json()['code'] == 0

        res = ac.get('/api/bill/classify/memory')
        assert res.get_json()['data'] == []


# ==================== 账单编辑：学习行为 ====================
class TestEditLearningRegression:
    """回归：用户编辑时改了分类 → 应该学习，让智能分类更准

    之前 bug：edit_bill 直接 UPDATE，没记用户修正 → 同一个「某某小店」
    永远分类错误。现在：old_record[4] != category 时调 remember_correction。
    """

    def test_edit_same_category_does_not_learn(self, reg_user):
        ac, uid = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '餐饮', 'note': '普通午餐'
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']

        # 编辑时分类没变
        ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 60, 'type': 'expense',
            'category': '餐饮', 'note': '普通午餐'
        })

        # 不应写入记忆
        res = ac.get('/api/bill/classify/memory')
        assert res.get_json()['data'] == []

    def test_edit_change_category_learns(self, reg_user):
        ac, uid = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '其他', 'note': '学习关键词A'
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']

        ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '餐饮', 'note': '学习关键词A'
        })

        # 后续分类应该走用户记忆
        res = ac.post('/api/bill/classify', json={'text': '学习关键词A'})
        d = res.get_json()['data']
        assert d['category'] == '餐饮'
        assert d.get('source') == 'user_memory'

    def test_edit_with_empty_note_does_not_learn(self, reg_user):
        """空 note 不学习（没东西可学）"""
        ac, uid = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '其他', 'note': ''
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']

        ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense',
            'category': '餐饮', 'note': ''
        })

        res = ac.get('/api/bill/classify/memory')
        assert res.get_json()['data'] == []

    def test_edit_missing_amount_rejected(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense', 'category': '餐饮'
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']

        res = ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2005

    def test_edit_invalid_amount_format_rejected(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense', 'category': '餐饮'
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']

        res = ac.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 'not_a_number', 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2006

    def test_edit_other_users_record_silent_fail(self, client):
        """A 创建的记录，B 编辑应该静默失败（不抛异常，DB 不变）"""
        suffix = str(int(time.time() * 1000))[-8:]
        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'edit_a_{suffix}', 'x'))
            uid_a = db.c.lastrowid
            db.conn.commit()
        add_record('2026-06-01', 100, 'expense', '餐饮', 'A的记录', uid_a)
        rid = get_records(uid_a)[0][0]

        # B 登录
        _, token, uid_b = _make_user()
        from gerenjizhang.utils.decorators import remove_token
        res = client.post(f'/api/bill/edit/{rid}', json={
            'date': '2026-06-01', 'amount': 999, 'type': 'expense', 'category': '购物', 'note': 'B改的'
        }, headers={'Authorization': f'Bearer {token}'})
        assert res.get_json()['code'] == 0  # 不报错（edit_record 不返回错误）

        # A 的记录金额应该没变
        from gerenjizhang.db import get_record_by_id
        rec = get_record_by_id(rid, uid_a)
        assert float(rec[2]) == 100
        assert rec[4] == '餐饮'

        # 清理
        remove_token(token)
        with Database() as db:
            db.c.execute('DELETE FROM bill WHERE user_id IN (%s, %s)', (uid_a, uid_b))
            db.c.execute('DELETE FROM users WHERE id IN (%s, %s)', (uid_a, uid_b))
            db.conn.commit()


# ==================== add_bill API 边界 ====================
class TestAddBillEdge:
    def test_add_bill_amount_zero_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 0, 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2005

    def test_add_bill_amount_negative_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': -50, 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2005

    def test_add_bill_invalid_type_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'transfer', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2004

    def test_add_bill_amount_string_coerced(self, reg_user):
        """字符串数字应该被 float() 转换"""
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': '99.50', 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 0
        rid = res.get_json()['data']['id']
        recs = get_records(ac.user_id)
        assert float([r for r in recs if r[0] == rid][0][2]) == 99.50

    def test_add_bill_amount_garbage_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 'not_a_number', 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2003

    def test_add_bill_missing_field(self, reg_user):
        ac, _ = reg_user
        for missing in ['date', 'type', 'category']:
            payload = {'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'}
            payload.pop(missing)
            res = ac.post('/api/bill/add', json=payload)
            assert res.get_json()['code'] == 2002, f"缺 {missing} 应返回 2002"


# ==================== 账单列表：分页 / 过滤 ====================
class TestBillListFilters:
    def test_pagination(self, reg_user):
        ac, _ = reg_user
        for i in range(25):
            ac.post('/api/bill/add', json={
                'date': f'2026-06-{(i % 30) + 1:02d}', 'amount': 10,
                'type': 'expense', 'category': '餐饮'
            })

        res = ac.get('/api/bill/list?year=2026&month=6&page=1&page_size=10')
        d = res.get_json()['data']
        assert d['total'] == 25
        assert d['page'] == 1
        assert d['page_size'] == 10
        assert d['total_pages'] == 3
        assert len(d['list']) == 10

        res = ac.get('/api/bill/list?year=2026&month=6&page=3&page_size=10')
        assert len(res.get_json()['data']['list']) == 5  # 25 mod 10

    def test_type_filter(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 5000, 'type': 'income', 'category': '工资'
        })

        res = ac.get('/api/bill/list?year=2026&month=6&type=expense')
        d = res.get_json()['data']
        assert d['total'] == 1
        assert d['list'][0]['type'] == 'expense'

    def test_category_filter(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense', 'category': '交通'
        })

        res = ac.get('/api/bill/list?year=2026&month=6&category=交通')
        d = res.get_json()['data']
        assert d['total'] == 1
        assert d['list'][0]['category'] == '交通'

    def test_total_pages_zero_safe(self, reg_user):
        """空月份 page_size=0 不应崩"""
        ac, _ = reg_user
        res = ac.get('/api/bill/list?year=2026&month=6&page_size=0')
        # 即便 page_size=0，total_pages 应能算出（不会除零）
        d = res.get_json()['data']
        assert 'total_pages' in d


# ==================== 统计接口 ====================
class TestStatsEndpoints:
    def test_monthly_basic(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-02', 'amount': 5000, 'type': 'income', 'category': '工资'
        })

        res = ac.get('/api/stats/monthly?year=2026&month=6')
        d = res.get_json()['data']
        assert d['income'] == 5000
        assert d['expense'] == 100
        assert d['balance'] == 4900

    def test_monthly_default_uses_now(self, reg_user):
        """不传 year/month → 用当前月（不应崩）"""
        ac, _ = reg_user
        res = ac.get('/api/stats/monthly')
        assert res.get_json()['code'] == 0

    def test_category_stats_with_type_filter(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-02', 'amount': 5000, 'type': 'income', 'category': '工资'
        })

        # 只看支出
        res = ac.get('/api/stats/category?year=2026&month=6&type=expense')
        d = res.get_json()['data']
        assert len(d) == 1
        assert d[0]['category'] == '餐饮'
        assert d[0]['type'] == 'expense'

        # 只看收入
        res = ac.get('/api/stats/category?year=2026&month=6&type=income')
        d = res.get_json()['data']
        assert len(d) == 1
        assert d[0]['category'] == '工资'

    def test_category_stats_percent(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 300, 'type': 'expense', 'category': '购物'
        })
        res = ac.get('/api/stats/category?year=2026&month=6')
        d = res.get_json()['data']
        by_cat = {c['category']: c for c in d}
        assert by_cat['购物']['percent'] == 75.0
        assert by_cat['餐饮']['percent'] == 25.0

    def test_daily_groups_by_date(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 50, 'type': 'expense', 'category': '交通'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-02', 'amount': 200, 'type': 'expense', 'category': '购物'
        })

        res = ac.get('/api/stats/daily?start=2026-06-01&end=2026-07-01')
        d = res.get_json()['data']
        by_date = {x['date']: x for x in d}
        assert by_date['2026-06-01']['amount'] == 150
        assert by_date['2026-06-01']['count'] == 2
        assert by_date['2026-06-02']['amount'] == 200

    def test_top_records_orders_by_amount_desc(self, reg_user):
        ac, _ = reg_user
        for amt in [10, 500, 100, 50, 1000]:
            ac.post('/api/bill/add', json={
                'date': '2026-06-01', 'amount': amt, 'type': 'expense', 'category': '餐饮'
            })
        res = ac.get('/api/stats/top?start=2026-06-01&end=2026-07-01&limit=3')
        d = res.get_json()['data']
        assert len(d) == 3
        assert d[0]['amount'] == 1000
        assert d[1]['amount'] == 500
        assert d[2]['amount'] == 100

    def test_top_records_income_type(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 5000, 'type': 'income', 'category': '工资'
        })
        res = ac.get('/api/stats/top?start=2026-06-01&end=2026-07-01&type=income')
        d = res.get_json()['data']
        assert len(d) == 1
        assert d[0]['amount'] == 5000

    def test_trend_returns_n_months(self, reg_user):
        ac, _ = reg_user
        res = ac.get('/api/stats/trend?months=6')
        d = res.get_json()['data']
        assert len(d['months']) == 6
        assert len(d['income']) == 6
        assert len(d['expense']) == 6

    def test_trend_with_data(self, reg_user):
        """trend 接口在有数据时 by_category 字典排序（按总金额倒序）"""
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-02', 'amount': 500, 'type': 'expense', 'category': '购物'
        })

        res = ac.get('/api/stats/trend?months=2')
        d = res.get_json()['data']
        # 总额大的分类（购物=500）应排在前面
        if d['by_category']:
            cats = list(d['by_category'].keys())
            if '购物' in cats and '餐饮' in cats:
                assert cats.index('购物') < cats.index('餐饮')

    def test_summary_previous_period(self, reg_user):
        """summary 端点：当前 vs 上一周期"""
        ac, _ = reg_user
        # 当前周期
        ac.post('/api/bill/add', json={
            'date': '2026-06-15', 'amount': 200, 'type': 'expense', 'category': '餐饮'
        })
        # 上一周期
        ac.post('/api/bill/add', json={
            'date': '2026-05-15', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })

        res = ac.get('/api/stats/summary?start=2026-06-01&end=2026-07-01')
        d = res.get_json()['data']
        assert d['expense'] == 200
        # 上一周期有数据
        assert d['previous'] is not None
        assert d['previous']['expense'] == 100
        # 增长率 = (200-100)/100 = 100%
        assert d['expense_delta'] == 100.0

    def test_summary_no_previous_data(self, reg_user):
        """没有上一周期数据时：previous 应是零值 dict，delta 应为 None（避免除零）"""
        ac, _ = reg_user
        res = ac.get('/api/stats/summary?start=2026-06-01&end=2026-07-01')
        d = res.get_json()['data']
        # 上一周期没有数据：summary 仍返回 dict（值全 0），但 delta 应为 None
        assert d['previous'] is not None
        assert d['previous']['expense'] == 0
        assert d['previous']['income'] == 0
        assert d['expense_delta'] is None  # 避免除以 0
        assert d['income_delta'] is None

    def test_summary_previous_period_zero_to_nonzero(self, reg_user):
        """上一周期是 0 → 不应除零崩溃"""
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-15', 'amount': 200, 'type': 'expense', 'category': '餐饮'
        })
        res = ac.get('/api/stats/summary?start=2026-06-01&end=2026-07-01')
        d = res.get_json()['data']
        # 上一周期 0 → delta = None（避免除零）
        assert d['expense_delta'] is None


# ==================== 回收站：跨用户 / 边界 ====================
class TestTrashCrossUser:
    def test_restore_other_users_record_fails(self, client):
        """A 的回收站记录，B 还原应该失败"""
        suffix = str(int(time.time() * 1000))[-8:]
        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'trash_a_{suffix}', 'x'))
            uid_a = db.c.lastrowid
            db.conn.commit()
        add_record('2026-06-01', 100, 'expense', '餐饮', '', uid_a)
        rid = get_records(uid_a)[0][0]
        # 软删
        from gerenjizhang.db import delete_record
        delete_record(rid, uid_a)

        # B 试图还原
        _, token, uid_b = _make_user()
        from gerenjizhang.utils.decorators import remove_token
        res = client.post(f'/api/bill/restore/{rid}',
                          headers={'Authorization': f'Bearer {token}'})
        assert res.get_json()['code'] == 4001

        # A 的记录应该还在回收站
        from gerenjizhang.db import get_record_by_id
        rec = get_record_by_id(rid, uid_a)
        assert rec is not None

        # 清理
        remove_token(token)
        with Database() as db:
            db.c.execute('DELETE FROM bill WHERE user_id IN (%s, %s)', (uid_a, uid_b))
            db.c.execute('DELETE FROM users WHERE id IN (%s, %s)', (uid_a, uid_b))
            db.conn.commit()

    def test_purge_other_users_record_fails(self, client):
        suffix = str(int(time.time() * 1000))[-8:]
        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'purge_a_{suffix}', 'x'))
            uid_a = db.c.lastrowid
            db.conn.commit()
        add_record('2026-06-01', 100, 'expense', '餐饮', '', uid_a)
        rid = get_records(uid_a)[0][0]
        from gerenjizhang.db import delete_record
        delete_record(rid, uid_a)

        _, token, uid_b = _make_user()
        from gerenjizhang.utils.decorators import remove_token
        res = client.post(f'/api/bill/purge/{rid}',
                          headers={'Authorization': f'Bearer {token}'})
        assert res.get_json()['code'] == 4002

        # A 的记录应该还在
        from gerenjizhang.db import get_record_by_id
        assert get_record_by_id(rid, uid_a) is not None

        remove_token(token)
        with Database() as db:
            db.c.execute('DELETE FROM bill WHERE user_id IN (%s, %s)', (uid_a, uid_b))
            db.c.execute('DELETE FROM users WHERE id IN (%s, %s)', (uid_a, uid_b))
            db.conn.commit()

    def test_batch_delete_only_own_records(self, reg_user):
        """批量删除 ids 中混着别人的 id → 只能删自己的"""
        ac, uid = reg_user
        suffix = str(int(time.time() * 1000))[-8:]
        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'batch_other_{suffix}', 'x'))
            uid_other = db.c.lastrowid
            db.conn.commit()
        # 自己 1 条 + 别人 1 条
        add_record('2026-06-01', 100, 'expense', '餐饮', '', uid)
        add_record('2026-06-01', 200, 'expense', '交通', '', uid_other)
        my_rid = get_records(uid)[0][0]
        other_rid = get_records(uid_other)[0][0]

        # 把自己和别人的 id 一起发
        res = ac.post('/api/bill/batch-delete', json={'ids': [my_rid, other_rid]})
        assert res.get_json()['code'] == 0
        # 只删了 1 条（自己的），别人的 0 条
        assert res.get_json()['data']['deleted'] == 1

        # 别人的记录应该还在
        from gerenjizhang.db import get_record_by_id
        assert get_record_by_id(other_rid, uid_other) is not None

        with Database() as db:
            db.c.execute('DELETE FROM bill WHERE user_id=%s', (uid_other,))
            db.c.execute('DELETE FROM users WHERE id=%s', (uid_other,))
            db.conn.commit()

    def test_restore_batch_empty_ids(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/trash/restore-batch', json={'ids': []})
        assert res.get_json()['code'] == 4003

    def test_restore_batch(self, reg_user):
        ac, uid = reg_user
        ids = []
        for i in range(3):
            r = ac.post('/api/bill/add', json={
                'date': f'2026-06-0{i+1}', 'amount': 10, 'type': 'expense', 'category': '餐饮'
            })
            rid = r.get_json()['data']['id']
            ids.append(rid)
            ac.post(f'/api/bill/delete/{rid}')

        res = ac.post('/api/bill/trash/restore-batch', json={'ids': ids})
        assert res.get_json()['code'] == 0
        assert res.get_json()['data']['restored'] == 3

        # 列表里 3 条都回来了
        res = ac.get('/api/bill/list?year=2026&month=6')
        assert res.get_json()['data']['total'] == 3


# ==================== 预算 ====================
class TestBudgetEdge:
    def test_budget_string_amount_coerced(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': '500.50', 'month': '2026-06'
        })
        assert res.get_json()['code'] == 0

        res = ac.get('/api/bill/budget?year=2026&month=6')
        assert res.get_json()['data'][0]['budget'] == 500.50

    def test_budget_invalid_amount_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': 'garbage', 'month': '2026-06'
        })
        assert res.get_json()['code'] == 3002

    def test_budget_zero_allowed(self, reg_user):
        """预算 0（元）应该允许 → 等于关闭预算"""
        ac, _ = reg_user
        res = ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': 0, 'month': '2026-06'
        })
        assert res.get_json()['code'] == 0

    def test_budget_delete_nonexistent(self, reg_user):
        """删除不存在的预算不应该崩"""
        ac, _ = reg_user
        res = ac.post('/api/bill/delete-budget', json={
            'category': '不存在的分类', 'month': '2026-06'
        })
        # 应该返回成功（DELETE 是幂等的）
        assert res.get_json()['code'] == 0

    def test_budget_delete_missing_category_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/delete-budget', json={'month': '2026-06'})
        assert res.get_json()['code'] == 3005

    def test_budget_percent_no_division_by_zero(self, reg_user):
        """预算=0 时 percent 不应崩"""
        ac, _ = reg_user
        ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': 0, 'month': '2026-06'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        res = ac.get('/api/bill/budget?year=2026&month=6')
        d = res.get_json()['data'][0]
        # 预算=0，percent 应为 0 而不是除零
        assert d['percent'] == 0
        assert d['remaining'] == 0

    def test_budget_default_month(self, reg_user):
        """不传 month → 默认当前月"""
        ac, _ = reg_user
        from datetime import datetime
        res = ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': 500
        })
        assert res.get_json()['code'] == 0

        month_str = f"{datetime.now().year}-{datetime.now().month:02d}"
        res = ac.get(f'/api/bill/budget?month={month_str}')
        assert len(res.get_json()['data']) == 1

    def test_budget_aggregate_spent_across_records(self, reg_user):
        """spent 应累加当月所有该分类的支出（不是只算第一笔）"""
        ac, _ = reg_user
        ac.post('/api/bill/budget', json={
            'category': '餐饮', 'amount': 1000, 'month': '2026-06'
        })
        for amt in [100, 200, 300]:
            ac.post('/api/bill/add', json={
                'date': '2026-06-01', 'amount': amt, 'type': 'expense', 'category': '餐饮'
            })

        res = ac.get('/api/bill/budget?year=2026&month=6')
        d = res.get_json()['data'][0]
        assert d['spent'] == 600
        assert d['percent'] == 60.0
        assert d['remaining'] == 400


# ==================== Auth 回归 ====================
class TestAuthRegression:
    def test_register_new_user(self, client):
        """POST /api/auth/register 创建用户"""
        suffix = str(int(time.time() * 1000))[-8:]
        username = f'newreg_{suffix}'

        res = client.post('/api/auth/register', json={
            'username': username, 'password': 'validpass123'
        })
        assert res.get_json()['code'] == 0

        # 清理
        with Database() as db:
            db.c.execute('DELETE FROM users WHERE username=%s', (username,))
            db.conn.commit()

    def test_register_duplicate_username_rejected(self, client):
        suffix = str(int(time.time() * 1000))[-8:]
        username = f'dup_{suffix}'
        client.post('/api/auth/register', json={
            'username': username, 'password': 'validpass123'
        })
        res = client.post('/api/auth/register', json={
            'username': username, 'password': 'validpass123'
        })
        assert res.get_json()['code'] == 1010  # 用户名已存在

        with Database() as db:
            db.c.execute('DELETE FROM users WHERE username=%s', (username,))
            db.conn.commit()

    def test_register_short_username_rejected(self, client):
        res = client.post('/api/auth/register', json={
            'username': 'ab', 'password': 'validpass123'
        })
        assert res.get_json()['code'] == 1009

    def test_register_short_password_rejected(self, client):
        res = client.post('/api/auth/register', json={
            'username': 'goodname', 'password': '12345'
        })
        assert res.get_json()['code'] == 1009

    def test_register_missing_field(self, client):
        res = client.post('/api/auth/register', json={'username': 'goodname'})
        assert res.get_json()['code'] == 1008

    def test_login_wrong_password(self, client):
        suffix = str(int(time.time() * 1000))[-8:]
        username = f'wp_{suffix}'
        client.post('/api/auth/register', json={
            'username': username, 'password': 'correctpass'
        })
        res = client.post('/api/auth/login', json={
            'username': username, 'password': 'wrongpass'
        })
        assert res.get_json()['code'] == 1007

        with Database() as db:
            db.c.execute('DELETE FROM users WHERE username=%s', (username,))
            db.conn.commit()

    def test_login_nonexistent_user(self, client):
        res = client.post('/api/auth/login', json={
            'username': 'no_such_user_xyz', 'password': 'whatever'
        })
        assert res.get_json()['code'] == 1007

    def test_login_missing_fields(self, client):
        res = client.post('/api/auth/login', json={'username': 'foo'})
        assert res.get_json()['code'] == 1006

    def test_login_success_returns_token(self, client):
        suffix = str(int(time.time() * 1000))[-8:]
        username = f'loginok_{suffix}'
        client.post('/api/auth/register', json={
            'username': username, 'password': 'validpass123'
        })
        res = client.post('/api/auth/login', json={
            'username': username, 'password': 'validpass123'
        })
        assert res.get_json()['code'] == 0
        d = res.get_json()['data']
        assert 'token' in d
        assert d['userinfo']['username'] == username

        # 用 token 访问受保护接口
        res = client.get('/api/bill/list', headers={
            'Authorization': f'Bearer {d["token"]}'
        })
        assert res.get_json()['code'] == 0

        # 清理
        with Database() as db:
            db.c.execute('DELETE FROM tokens WHERE user_id IN (SELECT id FROM users WHERE username=%s)', (username,))
            db.c.execute('DELETE FROM users WHERE username=%s', (username,))
            db.conn.commit()

    def test_logout_invalidates_token(self, client):
        """登出后 token 失效"""
        suffix = str(int(time.time() * 1000))[-8:]
        username = f'logout_{suffix}'
        client.post('/api/auth/register', json={
            'username': username, 'password': 'validpass123'
        })
        res = client.post('/api/auth/login', json={
            'username': username, 'password': 'validpass123'
        })
        token = res.get_json()['data']['token']

        # 登出
        res = client.post('/api/auth/logout', headers={
            'Authorization': f'Bearer {token}'
        })
        assert res.get_json()['code'] == 0

        # 再访问应该 401
        res = client.get('/api/bill/list', headers={
            'Authorization': f'Bearer {token}'
        })
        assert res.status_code == 401

        # 清理
        with Database() as db:
            db.c.execute('DELETE FROM users WHERE username=%s', (username,))
            db.conn.commit()

    def test_get_userinfo(self, reg_user):
        ac, uid = reg_user
        res = ac.get('/api/auth/userinfo')
        assert res.get_json()['code'] == 0
        d = res.get_json()['data']
        assert d['id'] == uid


# ==================== 导入（CSV） ====================
class TestImportCSV:
    def test_import_basic_csv(self, reg_user):
        ac, uid = reg_user
        csv_content = "日期,类型,分类,金额,备注\n2026-06-01,expense,餐饮,100,午饭\n2026-06-02,income,工资,5000,月薪\n"
        res = ac.post('/api/bill/import', data={
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test.csv')
        }, content_type='multipart/form-data')
        d = res.get_json()
        assert d['code'] == 0, f"导入失败: {d}"
        assert d['data']['imported'] == 2

        # DB 里能查到
        recs = get_records(uid)
        assert len(recs) == 2

    def test_import_dedup(self, reg_user):
        """重复导入应该去重"""
        ac, uid = reg_user
        csv_content = "日期,类型,分类,金额,备注\n2026-06-01,expense,餐饮,100,午饭\n"
        # 第一次
        ac.post('/api/bill/import', data={
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test.csv')
        }, content_type='multipart/form-data')
        # 第二次相同
        res = ac.post('/api/bill/import', data={
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test.csv')
        }, content_type='multipart/form-data')
        d = res.get_json()
        assert d['code'] == 0
        assert d['data']['imported'] == 0
        assert d['data']['duplicate'] == 1

    def test_import_dry_run(self, reg_user):
        ac, uid = reg_user
        csv_content = "日期,类型,分类,金额,备注\n2026-06-01,expense,餐饮,100,午饭\n"
        res = ac.post('/api/bill/import', data={
            'file': (io.BytesIO(csv_content.encode('utf-8')), 'test.csv'),
            'dry_run': 'true'
        }, content_type='multipart/form-data')
        d = res.get_json()
        assert d['code'] == 0
        assert d['data']['total'] == 1
        assert len(d['data']['preview']) == 1
        # dry_run 不应真正导入
        assert len(get_records(uid)) == 0

    def test_import_no_file(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/import', data={}, content_type='multipart/form-data')
        assert res.get_json()['code'] == 4001

    def test_import_unsupported_format(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/import', data={
            'file': (io.BytesIO(b'hello'), 'test.txt')
        }, content_type='multipart/form-data')
        assert res.get_json()['code'] == 4003


# ==================== 导入（XLSX 微信格式） ====================
class TestImportXLSX:
    def _build_wechat_xlsx(self):
        """构造一个微信 XLSX 文件（含说明行 + 表头 + 2 条数据）"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # 模拟微信账单：前 16 行是说明
        for i in range(16):
            ws.append([f'说明行 {i}'])
        # 表头
        ws.append(['交易时间', '交易类型', '交易对方', '商品', '收/支', '金额(元)', '支付方式', '当前状态', '交易单号'])
        # 数据
        ws.append(['2026-06-01 12:00:00', '商户消费', '美团点评', '外卖', '支出', '28.50', '微信支付', '支付成功', 'WX001'])
        ws.append(['2026-06-02 09:30:00', '商户消费', '滴滴出行', '打车', '支出', '25.00', '微信支付', '支付成功', 'WX002'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_import_wechat_xlsx(self, reg_user):
        ac, uid = reg_user
        buf = self._build_wechat_xlsx()
        res = ac.post('/api/bill/import', data={
            'file': (buf, 'wechat.xlsx')
        }, content_type='multipart/form-data')
        d = res.get_json()
        assert d['code'] == 0, f"导入失败: {d}"
        # 2 条记录
        recs = get_records(uid)
        assert len(recs) == 2
        # 微信：交易对方是"美团点评" → 分类"餐饮"
        cats = {r[4] for r in recs}
        assert '餐饮' in cats
        assert '交通' in cats

    def test_import_wechat_skip_refund(self, reg_user):
        """'已退款' 状态应该跳过，不导入"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(16):
            ws.append([f'说明行 {i}'])
        ws.append(['交易时间', '交易类型', '交易对方', '商品', '收/支', '金额(元)', '支付方式', '当前状态', '交易单号'])
        ws.append(['2026-06-01 12:00:00', '商户消费', '美团点评', '外卖', '支出', '28.50', '微信支付', '已退款', 'WX001'])
        ws.append(['2026-06-02 09:30:00', '商户消费', '滴滴出行', '打车', '支出', '25.00', '微信支付', '支付成功', 'WX002'])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        ac, uid = reg_user
        res = ac.post('/api/bill/import', data={
            'file': (buf, 'wechat.xlsx')
        }, content_type='multipart/form-data')
        # 第一条已退款跳过，第二条成功
        recs = get_records(uid)
        assert len(recs) == 1
        assert recs[0][4] == '交通'


# ==================== 导出 ====================
class TestExport:
    def test_export_csv(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮', 'note': '午饭'
        })
        res = ac.get('/api/bill/export?year=2026&month=6&format=csv')
        assert res.status_code == 200
        assert 'text/csv' in res.headers.get('Content-Type', '')
        # 解码
        body = res.data.decode('utf-8') if isinstance(res.data, bytes) else res.data
        assert '餐饮' in body
        assert '午饭' in body
        assert '100' in body

    def test_export_xlsx(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮', 'note': '午饭'
        })
        res = ac.get('/api/bill/export?year=2026&month=6&format=xlsx')
        assert res.status_code == 200
        ct = res.headers.get('Content-Type', '')
        assert 'spreadsheetml' in ct or 'octet-stream' in ct

        # 解析 xlsx
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(res.data))
        ws = wb.active
        rows = list(ws.values)
        # 1 行表头 + 1 行数据
        assert len(rows) == 2
        assert rows[1][3] == 100  # 金额

    def test_export_default_is_csv(self, reg_user):
        ac, _ = reg_user
        res = ac.get('/api/bill/export?year=2026&month=6')
        assert 'text/csv' in res.headers.get('Content-Type', '')

    def test_export_excludes_deleted(self, reg_user):
        """导出不应该包括已删除的记录"""
        ac, uid = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮', 'note': '不导出'
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']
        ac.post(f'/api/bill/delete/{rid}')

        res = ac.get('/api/bill/export?year=2026&month=6&format=csv')
        body = res.data.decode('utf-8') if isinstance(res.data, bytes) else res.data
        assert '不导出' not in body


# ==================== 越权 / 注入 / 边界 ====================
class TestSecurityRegression:
    def test_token_tampering_returns_401(self, client):
        """篡改 token → 401"""
        res = client.get('/api/bill/list', headers={
            'Authorization': 'Bearer totally-fake-token-12345'
        })
        assert res.status_code == 401

    def test_sql_injection_in_category_filter(self, reg_user):
        """SQL 注入防御：category 参数不应被拼接"""
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        # 试图注入
        res = ac.get("/api/bill/list?year=2026&month=6&category=' OR '1'='1")
        # 不应崩，不应返回所有记录
        assert res.get_json()['code'] == 0
        # 不应返回餐饮的记录（因为 category 是 SQL 注入字符串）
        assert res.get_json()['data']['total'] == 0

    def test_sql_injection_in_login(self, client):
        """登录的 username 不应被注入"""
        res = client.post('/api/auth/login', json={
            'username': "admin' OR '1'='1", 'password': 'x'
        })
        # 应该返回 1007（账号或密码错误），不是 200/0
        assert res.get_json()['code'] == 1007

    def test_xss_in_note_does_not_break(self, reg_user):
        """XSS payload 存在 note → 应能正常存储/取出"""
        ac, _ = reg_user
        payload = "<script>alert('xss')</script>"
        ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 100, 'type': 'expense', 'category': '餐饮', 'note': payload
        })
        rid = ac.get('/api/bill/list').get_json()['data']['list'][0]['id']
        res = ac.get(f'/api/bill/{rid}')
        assert res.get_json()['data']['note'] == payload

    def test_negative_amount_via_api_rejected(self, reg_user):
        ac, _ = reg_user
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': -0.01, 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 2005

    def test_very_large_amount_via_api(self, reg_user):
        ac, _ = reg_user
        # 9999999.99 是 DECIMAL(10,2) 的极限 → 应成功
        res = ac.post('/api/bill/add', json={
            'date': '2026-06-01', 'amount': 9999999.99, 'type': 'expense', 'category': '餐饮'
        })
        assert res.get_json()['code'] == 0


# ==================== 月份边界 ====================
class TestMonthBoundary:
    def test_december_rolls_to_next_year(self, reg_user):
        """12 月查询应该正确处理跨年"""
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-12-15', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-11-15', 'amount': 50, 'type': 'expense', 'category': '交通'
        })

        res = ac.get('/api/bill/list?year=2026&month=12')
        d = res.get_json()['data']
        assert d['total'] == 1
        assert d['list'][0]['date'] == '2026-12-15'

        # 12 月的范围应该是 2026-12-01 到 2027-01-01
        # 11 月数据不会泄漏进 12 月

    def test_january_starts_year(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-01-15', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2025-12-31', 'amount': 50, 'type': 'expense', 'category': '交通'
        })

        res = ac.get('/api/bill/list?year=2026&month=1')
        d = res.get_json()['data']
        assert d['total'] == 1

    def test_february_short_month(self, reg_user):
        ac, _ = reg_user
        ac.post('/api/bill/add', json={
            'date': '2026-02-28', 'amount': 100, 'type': 'expense', 'category': '餐饮'
        })
        ac.post('/api/bill/add', json={
            'date': '2026-03-02', 'amount': 50, 'type': 'expense', 'category': '交通'
        })

        res = ac.get('/api/bill/list?year=2026&month=2')
        d = res.get_json()['data']
        assert d['total'] == 1
        assert d['list'][0]['date'] == '2026-02-28'


# ==================== Health check ====================
class TestHealthCheck:
    def test_health(self, client):
        res = client.get('/api/health')
        assert res.status_code == 200
        assert res.get_json()['status'] == 'healthy'

    def test_root_requires_auth(self, client):
        res = client.get('/')
        assert res.status_code == 401


# ==================== Token 行为 ====================
class TestTokenLifecycle:
    def test_expired_token_rejected(self, client):
        """过期的 token → 401"""
        suffix = str(int(time.time() * 1000))[-8:]
        token = f'expired_{suffix}'

        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'exp_user_{suffix}', 'x'))
            uid = db.c.lastrowid
            db.c.execute('''
                INSERT INTO tokens (token, user_id, expires_at)
                VALUES (%s, %s, NOW() - INTERVAL 1 DAY)
            ''', (token, uid))
            db.conn.commit()

        res = client.get('/api/bill/list', headers={
            'Authorization': f'Bearer {token}'
        })
        assert res.status_code == 401

        # 清理
        with Database() as db:
            db.c.execute('DELETE FROM tokens WHERE user_id=%s', (uid,))
            db.c.execute('DELETE FROM users WHERE id=%s', (uid,))
            db.conn.commit()

    def test_active_token_extends_expiry(self, client):
        """活跃 token 滑动过期（访问后 expires_at 应被延长）"""
        suffix = str(int(time.time() * 1000))[-8:]
        token = f'slide_{suffix}'

        with Database() as db:
            db.c.execute('INSERT INTO users (username, password) VALUES (%s, %s)',
                         (f'slide_user_{suffix}', 'x'))
            uid = db.c.lastrowid
            # 设 1 天后过期（已经"快过期"）
            db.c.execute('''
                INSERT INTO tokens (token, user_id, expires_at)
                VALUES (%s, %s, NOW() + INTERVAL 1 DAY)
            ''', (token, uid))
            db.conn.commit()

        res = client.get('/api/bill/list', headers={
            'Authorization': f'Bearer {token}'
        })
        assert res.status_code == 200

        # 过期时间应被推后到 7 天后
        with Database() as db:
            db.c.execute('SELECT expires_at FROM tokens WHERE token=%s', (token,))
            row = db.c.fetchone()
        from datetime import datetime, timedelta
        assert row[0] > datetime.now() + timedelta(days=6)

        # 清理
        with Database() as db:
            db.c.execute('DELETE FROM tokens WHERE user_id=%s', (uid,))
            db.c.execute('DELETE FROM users WHERE id=%s', (uid,))
            db.conn.commit()
